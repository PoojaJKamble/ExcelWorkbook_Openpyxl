#import modules
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, fills

#Creat Workbook
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = 'Summary'
timestr = datetime.datetime.now()

#Set coulumns size
sheet1.column_dimensions['A'].width = 30
sheet1.column_dimensions['B'].width = 20
sheet1.column_dimensions['C'].width = 15
sheet1.column_dimensions['D'].width = 6
sheet1.column_dimensions['E'].width = 6
sheet1.column_dimensions['F'].width = 15
sheet1.column_dimensions['G'].width = 6
sheet1.column_dimensions['H'].width = 15
sheet1.column_dimensions['I'].width = 15
sheet1.column_dimensions['J'].width = 19
sheet1.column_dimensions['K'].width = 20
#Define border formats
thin_border = Border(left=Side(border_style='thin'),
                     right=Side(border_style='thin'),
                     top=Side(border_style='thin'),
                     bottom=Side(border_style='thin'),
                    )
#Define fill formating
fill_cell = PatternFill(fill_type=fills.FILL_SOLID, start_color='FFFFFF', end_color='FFFFFF')

#define size of table
row_num = 9
col_num = 3
#location of table
row_loc = 1
col_loc = 0

for i in range(row_loc,row_loc+row_num):
    for j in range(col_loc,col_num+col_loc):
        sheet1.cell(row=i+1,column=j+1).border=thin_border
        if i==row_loc:
            sheet1.cell(row=i+1,column=j+1).border=thin_border
            sheet1.cell(row=i+1,column=j+1).fill=fill_cell
        if i==row_loc+row_num-1:
            sheet1.cell(row=i+1,column=j+1).border=thin_border


#Billing Volume Table
sheet1["A2"] = "Description"
sheet1['A2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=1)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["B2"] = "Material ID"
sheet1['B2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=2)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1["C2"] = "Volume(TB)"
sheet1['C2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=3)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["C8"] = "Server Count"
sheet1['C8'].font = Font(bold=True)
cell = sheet1.cell(row= 8,column=3)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1.merge_cells('A1:C1')
font_style = sheet1.cell(row = 1, column = 1, value = 'Billing Volume').font
sheet1['A1'].font = Font(bold=True,size = "15")
cell = sheet1.cell(row=1,column=1)
cell.alignment = Alignment(horizontal='center', vertical='center')


#Define border formats
thin_border = Border(left=Side(border_style = 'thin'),
                     right=Side(border_style = 'thin'),
                     top=Side(border_style = 'thin'),
                     bottom=Side(border_style = 'thin'),
                      )
#Define fill formating
fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='FFFFFF',end_color='FFFFFF')

#define size of table
row_num = 22
col_num = 4
#location of table
row_loc = 1
col_loc = 5

for i in range(row_loc,row_loc+row_num):
    for j in range(col_loc,col_num+col_loc):
        sheet1.cell(row=i+1,column=j+1).border=thin_border
        if i==row_loc:
            sheet1.cell(row=i+1,column=j+1).border=thin_border
            sheet1.cell(row=i+1,column=j+1).fill=fill_cell
        if i==row_loc+row_num-1:
            sheet1.cell(row=i+1,column=j+1).border=thin_border


#Array Summary Table
sheet1["F2"] = "Array"
sheet1['F2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=6)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["G2"] = "Tier"
sheet1['G2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=7)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["H2"] = "Array Total(TB)"
sheet1['H2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=8)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["I2"] = "Allocated(TB)"
sheet1['I2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=9)
cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1.merge_cells('F1:I1')
font_style1 = sheet1.cell(row = 1, column = 6, value = 'Array Summary').font
sheet1['F1'].font = Font(bold=True,size = "15")
cell = sheet1.cell(row=1,column=6)
cell.alignment = Alignment(horizontal='center', vertical='center')



sheet1["L2"] = "Tier 1"
sheet1['L2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=12)
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["M2"] = "Tier 3"
sheet1['M2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=13)
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["N2"] = "Tier 4"
sheet1['N2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=14)
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["O2"] = "BUR"
sheet1['O2'].font = Font(bold=True)
cell = sheet1.cell(row= 2,column=15)
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1["K3"] = "Configured Capacity:"
cell = sheet1.cell(row= 3,column=11)
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1["K4"] = "50% Capacity:"
sheet1['K4'].font = Font(color='008000')
cell = sheet1.cell(row= 4,column=11)
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1["K5"] = "Sabre Forecast:"
sheet1['K5'].font = Font(color="ff0000")
cell = sheet1.cell(row= 5,column=11)
cell.alignment = Alignment(horizontal='center', vertical='center')


sheet1["K6"] = "Allocated:"
cell = sheet1.cell(row= 6,column=11)
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet1["K7"] = "Billing Volume:"
sheet1['K7'].font = Font(bold='True')
cell = sheet1.cell(row= 7,column=11)
cell.alignment = Alignment(horizontal='center', vertical='center')


#Set coulumns size for HP
sheet2 = workbook.create_sheet('HP')
sheet2.column_dimensions['A'].width = 13
sheet2.column_dimensions['B'].width = 13
sheet2.column_dimensions['C'].width = 75
sheet2.column_dimensions['D'].width = 13
sheet2.column_dimensions['E'].width = 13
sheet2.column_dimensions['F'].width = 200


sheet2['A1'].font = Font(bold=True)
sheet2.cell(row= 1,column=1,value="LUN#")

sheet2['B1'].font = Font(bold=True)
sheet2.cell(row= 1,column=2, value="VIR LUN ID")

sheet2['C1'].font = Font(bold=True)
sheet2.cell(row= 1,column=3,value="LUN NAME")

sheet2['D1'].font = Font(bold=True)
sheet2.cell(row= 1,column=4,value="Allocated")

sheet2['E1'].font = Font(bold=True)
sheet2.cell(row= 1,column=5,value="Used")

sheet2['F1'].font = Font(bold=True)
sheet2.cell(row= 1,column=6,value= "Port Info")

#Set coulumns size for NS
sheet3 = workbook.create_sheet('NS')
sheet3.column_dimensions['A'].width = 13
sheet3.column_dimensions['B'].width = 13
sheet3.column_dimensions['C'].width = 45
sheet3.column_dimensions['D'].width = 10
sheet3.column_dimensions['E'].width = 10
sheet3.column_dimensions['F'].width = 200

sheet3['A1'].font = Font(bold=True)
sheet3.cell(row= 1,column=1,value="LUN#")

sheet3['B1'].font = Font(bold=True)
sheet3.cell(row= 1,column=2, value="VIR LUN")

sheet3['C1'].font = Font(bold=True)
sheet3.cell(row= 1,column=3,value="LUN NAME")

sheet3['D1'].font = Font(bold=True)
sheet3.cell(row= 1,column=4,value="Allocated")

sheet3['E1'].font = Font(bold=True)
sheet3.cell(row= 1,column=5,value="Used")

sheet3['F1'].font = Font(bold=True)
sheet3.cell(row= 1,column=6,value= "Port Info")

#Set coulumns size for NL & BR
sheet4 = workbook.create_sheet('NL')
sheet4.column_dimensions['A'].width = 13
sheet4.column_dimensions['B'].width = 45
sheet4.column_dimensions['C'].width = 15
sheet4.column_dimensions['D'].width = 10


sheet4['A1'].font = Font(bold=True)
sheet4.cell(row= 1,column=1,value="Array")

sheet4['B1'].font = Font(bold=True)
sheet4.cell(row= 1,column=2, value="Export")

sheet4['C1'].font = Font(bold=True)
sheet4.cell(row= 1,column=3,value="Description")

sheet4['D1'].font = Font(bold=True)
sheet4.cell(row= 1,column=4,value="Allocated")

sheet4['E1'].font = Font(bold=True)
sheet4.cell(row= 1,column=5,value="Used")

#SAN Attached
sheet5 = workbook.create_sheet('SAN')
sheet5.column_dimensions['A'].width = 13
sheet5.column_dimensions['B'].width = 15
sheet5.column_dimensions['C'].width = 10

sheet5['A1'].font = Font(bold=True)
sheet5.cell(row= 1,column=1,value="Host Name")

sheet5['B1'].font = Font(bold=True)
sheet5.cell(row= 1,column=2, value="Array")

sheet5['C1'].font = Font(bold=True)
sheet5.cell(row= 1,column=3, value="LoggedIn")


#Save Workbook
workbook.save(filename='Billing-'+timestr.strftime('%d%b%Y')+'.xlsx')
print("Worksheet created Successfully!")


